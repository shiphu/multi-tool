from flask import abort
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.ext.automap import automap_base
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from config_generator import strip_whitespace
import ipaddress
from zipfile import ZipFile

# Initial configuration of SQLAlchemy engine and connection
DATABASE = 'sqlite:///database/dataloader_db.sqlite?check_same_thread=False'
Base = automap_base()
engine = create_engine(DATABASE)
Base.prepare(engine, reflect=True)

# Map tables from database
Equipment = Base.classes.equipment_templates
Srx345Ports = Base.classes.srx345_physical_ports
Asr9k = Base.classes.asr9k_routers
PeSites = Base.classes.pe_sites
FacilityVpnPorts = Base.classes.hf_vpn_ports
ServiceVpnPorts = Base.classes.service_vpn_ports
NmsPePorts = Base.classes.nms_pe_ports
ServicePePorts = Base.classes.service_pe_ports
SubnetMasks = Base.classes.subnet_masks

session = Session(engine)

# Folders and file names
DATA_FOLDER = Path('served_files/dataloader')
TEMPLATE_FILE = str(DATA_FOLDER / 'Template.xlsx')
R3_FILE = str(DATA_FOLDER / 'Region 3 Data File.xlsx')
R1_FILE = str(DATA_FOLDER / 'Region 1 Data File.xlsx')


class Shelf:
    def __init__(self, shelf_id, device_type, site_id, shelf_template=None, model=None, node='NODE 0'):
        self.shelf_id = shelf_id
        self.device_type = device_type
        self.site_id = site_id
        self.shelf_template = shelf_template
        self.node = node
        self.model = model
        
    def make_port(self, path_type, interface, vlan=None, network_addr=None, port_id=None, bandwidth=None):
        self.port = Port(self, path_type, interface, vlan, network_addr, port_id, bandwidth)
        return self.port


class Port():
    def __init__(self, device, path_type, interface, vlan=None, network_addr=None, port_id=None, bandwidth=None):
        self.shelf_id = device.shelf_id
        self.site_id = device.site_id
        self.path_type = path_type
        self.device_type = device.device_type
        self.node = device.node
        self.model = device.model
        self.interface = interface
        self.bandwidth = '{}M'.format(bandwidth) if bandwidth else 'UNDEFINED'
        self.connector = 'Logical'
        self.ip_addr = self.set_ip(network_addr)
        self.network = 'SSHA'
        self.customer = 'C001645'

        if self.device_type == 'pe' or self.device_type == 'vpn':
            table = self.get_table(path_type, self.device_type)
            q = session.query(table).filter(table.shelf_clei == self.shelf_id).first()
            self.port_id = self.make_port(q, interface, vlan) if not port_id else port_id
        elif self.device_type == 'cpe':
            if interface == 'loopback':
                q = session.query(Equipment).filter(Equipment.shelf_node == self.node).first()
                self.port_id = q.loopback_interface
            elif interface == 'logical':
                q = session.query(Equipment).filter(Equipment.shelf_node == self.node).first()
                self.port_id = '{}.{}'.format(port_id, vlan) if vlan else None
            elif interface == 'physical':
                q = session.query(Equipment).filter(Equipment.shelf_node == self.node).first()
                port_table = self.get_table(path_type, self.device_type, self.model)
                q = session.query(port_table).filter(port_table.physical_interface == port_id).first()
                self.port_id = port_id
        elif self.device_type == 'ts':
            q = session.query(Equipment).filter(Equipment.shelf_clei == self.model).first()
            self.port_id = port_id
            
        try:
            self.slot_1 = q.slot_1
            self.slot_2 = q.slot_2
            self.slot_3 = q.slot_3
        except AttributeError:
            pass

        self.hostname = '{} - {}'.format(self.shelf_id, self.port_id)
        
    def make_port(self, query, interface, vlan):
        if interface == 'logical' and vlan:
            self.port_id = '{}{}'.format(query.logical_interface, vlan)
        elif interface == 'physical' and vlan:
            self.port_id = '{}{}'.format(query.physical_interface, vlan)
        else:
            return None
        return self.port_id
        
    def get_table(self, path_type, device_type, model=None):
        tables = {
            'service_pe': ServicePePorts,
            'service_vpn': ServiceVpnPorts,
            'nms_pe': NmsPePorts,
            'facility_vpn': FacilityVpnPorts,
            'srx345_physical_ports': Srx345Ports
        }
        if device_type == 'vpn' or device_type =='pe':
            table = '{}_{}'.format(path_type, device_type)
        elif device_type == 'cpe':
            table = '{}_physical_ports'.format(model.lower())
        return tables[table]

    def set_ip(self, network_addr):
        ip_addr = ''
        if self.device_type == 'cpe':
            if self.interface == 'loopback':
                ip_addr = self.validate_ip(network_addr)
            else:
                ip_addr = self.validate_ip(network_addr, 1) 
        elif self.device_type == 'vpn':
            # Assign IP to logical port of primary VPN concentrator
            if self.path_type == 'facility' and self.interface == 'logical' and '45W' in self.shelf_id:
                ip_addr = self.validate_ip(network_addr)
            # Assign IP to physical port of primary VPN concentrator
            elif self.path_type == 'service' and self.interface == 'physical' and '45W' in self.shelf_id:
                ip_addr = self.validate_ip(network_addr, 2)
        elif self.device_type == 'pe':
            if self.path_type == 'nms':
                ip_addr = self.validate_ip(network_addr)
            elif self.path_type == 'service':
                ip_addr = self.validate_ip(network_addr, 1)
        elif self.device_type == 'ts':
            ip_addr = self.validate_ip(network_addr, 2)
        else: 
            return None
        
        return ip_addr
    
    def validate_ip(self, network_addr, increment=0):
        try:
            ip_addr = str(ipaddress.ip_address(network_addr) + int(increment))
        except ValueError:
            return None
        return ip_addr
    
    
#class Port():
#    def __init__(
#            self, shelf_id, device_type,
#            vlan=None, path_type=None, bandwidth=None,
#            interface=None, network_addr=None, port=None,
#            site_id=None, model=None, node=None):
#        super().__init__(shelf_id, site_id)
#        if device_type == 'pe' or device_type == 'vpn':
#            table = self.get_table(path_type, device_type)
#            q = session.query(table).filter(table.shelf_clei == self.shelf_id).first()
#            self.site_id = q.pop_clli
#            self.port_id = self.make_port(q, interface, vlan)
#        elif device_type == 'cpe':
#            if interface == 'loopback':
#                q = session.query(Equipment).filter(Equipment.shelf_node == self.node).first()
#                self.site_id = site_id
#                self.port_id = q.loopback_interface
#            elif interface == 'logical':
#                q = session.query(Equipment).filter(Equipment.shelf_node == self.node).first()
#                self.port_id = '{}.{}'.format(port, vlan) if vlan else None
#            elif interface == 'physical':
#                q = session.query(Equipment).filter(Equipment.shelf_node == self.node).first()
#                port_table = self.get_table(path_type, device_type, model)
#                q = session.query(port_table).filter(port_table.physical_interface == port).first()
#                self.site_id = site_id
#                self.port_id = port
#        elif device_type == 'ts':
#            q = session.query(Equipment).filter(Equipment.shelf_clei == model).first()
#            self.site_id = site_id
#            self.port_id = port
#            
#        try:
#            self.slot_1 = q.slot_1
#            self.slot_2 = q.slot_2
#            self.slot_3 = q.slot_3
#        except AttributeError:
#            pass
#        self.device_type = device_type
#        self.path_type = path_type
#        self.interface = interface
#        self.ip_addr = self.set_ip(network_addr) if network_addr else None
#        self.bandwidth = '{}M'.format(bandwidth) if bandwidth else 'UNDEFINED'
#        self.connector = 'Logical'
#        self.network = 'SSHA'
#        self.customer = 'C001645'
#        self.hostname = '{} - {}'.format(self.shelf_id, self.port_id)
#
#    def make_port(self, query, interface, vlan):
#        if interface == 'logical' and vlan:
#            self.port_id = '{}{}'.format(query.logical_interface, vlan)
#        elif interface == 'physical' and vlan:
#            self.port_id = '{}{}'.format(query.physical_interface, vlan)
#        else:
#            return None
#        return self.port_id



#    def set_ip(self, network_addr):
#        if self.path_type == 'facility' or self.path_type == 'nms':
#            # Assign IP to logical port of primary VPN concentrator
#            if self.device_type == 'vpn' and self.interface == 'logical' and '45W' in self.shelf_id:
#                ip_addr = self.validate_ip(network_addr)
#            elif self.device_type == 'pe':
#                ip_addr = self.validate_ip(network_addr)
#            elif self.device_type == 'cpe':
#                ip_addr = self.validate_ip(network_addr, 1)
#        elif self.path_type == 'service':
#            if self.device_type == 'pe':
#                ip_addr = self.validate_ip(network_addr, 1)
#            elif self.device_type == 'vpn' and self.interface == 'physical' and '45W' in self.shelf_id:
#                ip_addr = self.validate_ip(network_addr, 2)
#            elif self.device_type == 'cpe':
#                ip_addr = self.validate_ip(network_addr, 1) 
#        elif self.interface == 'loopback':
#            ip_addr = self.validate_ip(network_addr)
#        else: 
#            return None
#        ip_addr = ''
#        if self.device_type == 'cpe':
#            if self.interface == 'loopback':
#                ip_addr = self.validate_ip(network_addr)
#            else:
#                ip_addr = self.validate_ip(network_addr, 1) 
#        elif self.device_type == 'vpn':
#            # Assign IP to logical port of primary VPN concentrator
#            if self.path_type == 'facility' and self.interface == 'logical' and '45W' in self.shelf_id:
#                ip_addr = self.validate_ip(network_addr)
#            # Assign IP to physical port of primary VPN concentrator
#            elif self.path_type == 'service' and self.interface == 'physical' and '45W' in self.shelf_id:
#                ip_addr = self.validate_ip(network_addr, 2)
#        elif self.device_type == 'pe':
#            if self.path_type == 'nms':
#                ip_addr = self.validate_ip(network_addr)
#            elif self.path_type == 'service':
#                ip_addr = self.validate_ip(network_addr, 1)
#        elif self.device_type == 'ts':
#            ip_addr = self.validate_ip(network_addr, 2)
#        else: 
#            return None
#        
#        return ip_addr
#
#
#    def validate_ip(self, network_addr, increment=0):
#        try:
#            ip_addr = str(ipaddress.ip_address(network_addr) + int(increment))
#        except ValueError:
#            return None
#        return ip_addr


class Subnet:
    def __init__(self, network_addr, cidr):
        self.network = 'SSHA'
        self.sub_start = network_addr
        self.sub_end = ''
        if str(cidr) == '31':
            self.cidr = ''
            try:
                self.sub_end = str(ipaddress.ip_address(network_addr) + 1)
            except ValueError:
                pass
        else:
            self.cidr = cidr
        self.status = 'ACTIVE'
        self.notes = 'SSHA'


class Channel:
    def __init__(self, shelf_id, path_type, vlan, bandwidth=None, vendor=None):
        q = session.query(PeSites).filter(PeSites.shelf_clei == shelf_id).first()
        if path_type == 'service':
            self.path_name = q.pe_vpn_path
        elif path_type == 'facility':
            self.path_name = q.facility_path
        elif path_type == 'nms':
            self.path_name = q.nms_path
        elif path_type == 'vendor':
            self.path_name = vendor
        else:
            self.path_name = None
        
        self.channel = 'VLAN{}'.format(vlan) if vlan else None
        self.bandwidth = '{}M'.format(bandwidth) if bandwidth else 'UNDEFINED'
        self.status = 'Ok'


def create_r3_data_file(form, site_type):
    # Elements will be created and added into their respective key/value pair, then further validated and added to print_rows
    rows = {
        'shelf': [],
        'port': [],
        'ip': [],
        'subnet': [],
        'channel': []
    }
    
    # Validate all objects from rows dict and append ones that meet criteria for printing to sheet
    print_rows = {
        'shelf': [],
        'port': [],
        'ip': [],
        'subnet': [],
        'channel': []
    }
    
    d = strip_whitespace(form)
    site_suffix = ['', 'Backup'] if site_type == 'dual' else ['']

    cpe_router = Shelf(
        shelf_id=d['cpeRouter'],
        device_type='cpe', 
        site_id=d['siteCLLI'], 
        shelf_template=d['cpeTemplate'],
        model=d['modelName']
    )
    rows['shelf'].append(cpe_router)
    
    for site in site_suffix:
        pe = d['peRouter1']
        reth = d['rethInterface']
        bw_minus_one = int(d['siteBandwidth']) - 1
        if site == 'Backup':
            pe = d['peRouter2']
            reth = 'RETH20'
            bw_minus_one = int(d['siteBandwidthBackup']) - 1

        q = session.query(PeSites).filter(PeSites.shelf_clei == pe).first()

        pe_router = Shelf(shelf_id=pe, device_type='pe', site_id=q.pop_clli)
        primary_vpn = Shelf(shelf_id=q.vpn_primary, device_type='vpn', site_id=q.pop_clli)
        backup_vpn = Shelf(shelf_id=q.vpn_backup, device_type='vpn', site_id=q.pop_clli)
        vpn_concentrators = [primary_vpn, backup_vpn]
        interfaces = ['physical', 'logical']

        # NMS CPE & PE ports
        rows['port'].extend([
            pe_router.make_port(
                path_type='nms',
                interface='logical',
                vlan=d['nmsVlan{}'.format(site)],
                network_addr=d['nmsIP{}'.format(site)],
                bandwidth='1'
            ),
            cpe_router.make_port(
                path_type='nms',
                interface='logical',
                vlan=d['nmsVlan{}'.format(site)],
                network_addr=d['nmsIP{}'.format(site)],
                port_id=reth,
                bandwidth='1',
            ),
            # HOT Facility CPE port
            cpe_router.make_port(
                path_type='facility',
                interface='logical',
                vlan=d['facilityVlan{}'.format(site)],
                network_addr=d['facilityIP{}'.format(site)],
                port_id=reth,
                bandwidth=bw_minus_one
            )
        ])

        # HOT Facility VPN ports
        for vpn in vpn_concentrators:
            for interface in interfaces:
                rows['port'].append(vpn.make_port(
                    path_type='facility',
                    interface=interface,
                    vlan=d['facilityVlan{}'.format(site)],
                    network_addr=d['facilityIP{}'.format(site)],
                    bandwidth=bw_minus_one
                ))

        rows['subnet'].extend([
            Subnet(d['nmsIP{}'.format(site)], cidr='31'),
            Subnet(d['facilityIP'.format(site)], cidr='31')
        ])

        counter = 1  # Counter for suffix of HTML element names
        for i in range(int(d['numServices'])):
            suffix = '' if counter == 1 else counter
            rows['port'].append(pe_router.make_port(
                path_type='service',
                interface='logical',
                vlan=d['serviceVlan{}{}'.format(site, suffix)],
                network_addr=d['servicePeIP{}{}'.format(site, suffix)]
            ))
            for vpn in vpn_concentrators:
                for interface in interfaces:
                    rows['port'].append(vpn.make_port(
                        path_type='service',
                        interface=interface,
                        vlan=d['serviceVlan{}{}'.format(site, suffix)],
                        network_addr=d['servicePeIP{}{}'.format(site, suffix)]
                    ))

            rows['subnet'].append(Subnet(network_addr=d['servicePeIP{}{}'.format(site, suffix)], cidr='30'))

            counter += 1

    # Repeat service loop for CPE ports (only need to be created for primary)
    counter = 1  # Suffix of HTML element names
    for i in range(int(d['numServices'])):
        suffix = '' if counter == 1 else counter
        rows['ip'].append(cpe_router.make_port(
            path_type='service',
            interface='physical',
            network_addr=d['serviceCpeIP{}'.format(suffix)],
            port_id=d['servicePort{}'.format(suffix)]
        ))

        rows['subnet'].append(Subnet(
            network_addr=d['serviceCpeIP{}'.format(suffix)],
            cidr=d['serviceCpeCIDR{}'.format(suffix)]
        ))

        counter += 1
    for row in rows['port']:
        if row.port_id and row.shelf_id:
            print_rows['port'].append(row)
            if row.ip_addr:
                print_rows['ip'].append(row)
    for row in rows['ip']:
        if row.ip_addr and row.port_id and row.shelf_id:
            print_rows['ip'].append(row)
    for row in rows['subnet']:
        if row.sub_start:
            print_rows['subnet'].append(row)

            
    wb = load_workbook(TEMPLATE_FILE)  
    wb.save(R3_FILE)
    for sheet, row in print_rows.items():
        print_sheet(wb, R3_FILE, sheet, row)
            
def print_sheet(wb, file, sheet, rows):

    sheets = {
        'shelf': wb['SHELF'],
        'port': wb['PORT'],
        'subnet': wb['SUBNET'],
        'ip': wb['IP'],
        'channel': wb['CHANNEL'],
    }

    # Attributes ordered by columns according to Excel template
    sheet_attributes = {
        'shelf': ['shelf_id', 'site_id', 'shelf_template'],
        'port': [
            'shelf_id', 'site_id', 'slot_1',
            'slot_2', 'slot_3', 'port_id',
            'bandwidth', 'connector', 'ip_addr'],
        'subnet': [
            'network', 'sub_start', 'cidr',
            'sub_end', 'status', 'notes'],
        'ip': [
            'shelf_id', 'site_id', 'slot_1',
            'slot_2', 'slot_3', 'port_id',
            'ip_addr', 'network', 'customer',
            'hostname'],
        'channel': [
            'path_name', 'channel', 'bandwidth',
            'status']
    }

    active_sheet = sheets[sheet]
    attributes = sheet_attributes[sheet]

    row = 2
    for r in rows:
        col = 1
        for attr in attributes:
            if getattr(r, attr) != '':
                active_sheet['{}{}'.format(get_column_letter(col), row)] = str(getattr(r, attr))
            col += 1
        row += 1

    wb.save(file)


def dropdown_lists():
    lists = {
        'pe_routers': [i.shelf_clei for i in session.query(PeSites.shelf_clei)],
        'router_templates': [i.template_name for i in session.query(Equipment.template_name)
                             .filter(Equipment.equipment_type == 'ROUTER')
                             ],
        'server_templates': [i.template_name for i in session.query(Equipment.template_name)
                             .filter(Equipment.equipment_type == 'SERVER')
                             ],
        'cidr_prefixes': list(reversed([i.cidr_prefix for i in session.query(SubnetMasks.cidr_prefix)])),
        'srx345_node0_ports': [i.physical_interface for i in session.query(Srx345Ports.physical_interface)
                               .filter(Srx345Ports.shelf_node == 'NODE 0')
                               ]
    }
    # Prevent duplicate router models from being printed to list
    model_names = []
    for i in session.query(Equipment).filter(Equipment.equipment_type == 'ROUTER'):
        if i.shelf_clei not in model_names:
            model_names.append(i.shelf_clei)
    lists.update({'model_names': model_names})

    return lists


if __name__ == '__main__':
    form = {
         'siteType': 'single',
         'modelName': 'SRX345',
         'numServices': '3',
         'siteCLLI': 'clei',
         'peRouter1': 'TOROONXNPED10',
         'siteBandwidth': '100',
         'siteBandwidthBackup': '',
         'cpeTemplate': 'JUNIPER SRX345 SINGLE CPE',
         'cpeRouter': 'cpe clei',
         'cpeLoopback': '127.0.0.1',
        'rethInterface': 'RETH0',
        'terServer': '',
        'terServerIP1': '',
        'terServerCIDR1': '30',
        'terServerIP2': '',
        'terServerCIDR2': '30',
        'nmsVlan': '600',
        'nmsIP': '0.0.0.0',
        'nmsCIDR': '31',
        'facilityVlan': '601',
        'facilityIP': '20.0.0.0',
        'facilityCIDR': '31',
        'vendorVlan': '',
        'vendorChannel': '',
        'vendorPort': '',
        'nmsVlanBackup': '',
        'nmsIPBackup': '',
        'nmsCIDRBackup': '31',
        'facilityVlanBackup': '',
        'facilityIPBackup': '',
        'facilityCIDRBackup': '31',
        'vendorVlanBackup': '',
        'vendorChannelBackup': '',
        'vendorPortBackup': '',
        'serviceVlan': '100',
        'servicePort': 'GE-0/0/8',
        'serviceCpeIP': '0.0.0.0',
        'serviceCpeCIDR': '30',
        'servicePeIP': '7.0.0.0',
        'servicePeCIDR': '30',
        'serviceVlan2': '101',
        'servicePort2': 'GE-0/0/9',
        'serviceCpeIP2': '1.0.0.0',
        'serviceCpeCIDR2': '30',
        'servicePeIP2': '8.0.0.0',
        'servicePeCIDR2': '30',
        'serviceVlan3': '102',
        'servicePort3': 'GE-0/0/10',
        'serviceCpeIP3': '2.0.0.0',
        'serviceCpeCIDR3': '30',
        'servicePeIP3': '9.0.0.0',
        'servicePeCIDR3': '30',
        'serviceVlanBackup': '',
        'servicePeIPBackup': '',
        'servicePeCIDRBackup': '30'}

    
    create_r3_data_file(form, 'single')

    
    pass