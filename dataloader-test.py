from flask import abort
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.ext.automap import automap_base
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from config_generator import strip_whitespace
import ipaddress
from zipfile import ZipFile

# Initial configuration of SQLAlchemy engine and connection
DATABASE = 'sqlite:///database/dataloader_db.sqlite?check_same_thread=False'
Base = automap_base()
engine = create_engine(DATABASE)
Base.prepare(engine, reflect=True)

# Map tables from database
Equipment = Base.classes.equipment_templates
Srx345Ports = Base.classes.srx345_physical_ports
Asr9k = Base.classes.asr9k_routers
PeSites = Base.classes.pe_sites
FacilityVpnPorts = Base.classes.hf_vpn_ports
ServiceVpnPorts = Base.classes.service_vpn_ports
NmsPePorts = Base.classes.nms_pe_ports
ServicePePorts = Base.classes.service_pe_ports
SubnetMasks = Base.classes.subnet_masks

session = Session(engine)

# Folders and file names
DATA_FOLDER = Path('served_files/dataloader')
TEMPLATE_FILE = str(DATA_FOLDER / 'Template.xlsx')
R3_FILE = str(DATA_FOLDER / 'Region 3 Data File.xlsx')
R1_FILE = str(DATA_FOLDER / 'Region 1 Data File.xlsx')


class Shelf:
    def __init__(self, shelf_id, site_id=None, shelf_template=None, ip_addr=None, node='NODE 0'):
        self.shelf_id = shelf_id
        self.site_id = site_id
        self.shelf_template = shelf_template
        self.ip_addr = ip_addr
        self.node = node


class Port(Shelf):
    def __init__(
            self, shelf_id, device_type,
            path_type, vlan, bandwidth=None,
            interface=None, network_addr=None, port=None,
            site_id=None):
        super().__init__(shelf_id, site_id)
        if device_type == 'pe' or device_type == 'vpn':
            table = self.get_table(path_type, device_type)
            q = session.query(table).filter(table.shelf_clei == self.shelf_id).first()
            self.site_id = q.pop_clli
            self.port_id = self.make_port(q, interface, vlan)

        elif device_type == 'cpe':
            q = session.query(Equipment).filter(Equipment.shelf_node == self.node).first()
            self.port_id = '{}.{}'.format(port, vlan)

        self.slot_1 = q.slot_1
        self.slot_2 = q.slot_2
        self.slot_3 = q.slot_3
        self.device_type = device_type
        self.path_type = path_type
        self.interface = interface
        self.ip_addr = self.set_ip(network_addr) if network_addr else None
        self.bandwidth = '{}M'.format(bandwidth) if bandwidth else 'UNDEFINED'
        self.connector = 'LOGICAL'
        self.network = 'SSHA'
        self.customer = 'C001645'
        self.hostname = '{} - {}'.format(self.shelf_id, self.port_id)

    def make_port(self, query, interface, vlan):
        if interface == 'logical':
            self.port_id = '{}{}'.format(query.logical_interface, vlan)
        elif interface == 'physical':
            self.port_id = '{}{}'.format(query.physical_interface, vlan)
        return self.port_id

    def get_table(self, port_type, device_type):
        tables = {
            'service_pe': ServicePePorts,
            'service_vpn': ServiceVpnPorts,
            'nms_pe': NmsPePorts,
            'facility_vpn': FacilityVpnPorts,
        }
        table = '{}_{}'.format(port_type, device_type)
        return tables[table]

    def set_ip(self, network_addr):
        ip_addr = ''

        if self.path_type == 'facility' or self.path_type == 'nms':
            # Assign IP to logical port of primary VPN concentrator
            if self.device_type == 'vpn' and self.interface == 'logical' and '45W' in self.shelf_id:
                ip_addr = self.validate_ip(network_addr)
            elif self.device_type == 'cpe':
                ip_addr = self.validate_ip(network_addr, 1)
        elif self.path_type == 'service':
            if self.device_type == 'pe':
                ip_addr = self.validate_ip(network_addr, 1)
            elif self.device_type == 'vpn' and self.interface == 'physical' and '45W' in self.shelf_id:
                ip_addr = self.validate_ip(network_addr, 2)

        else:
            ip_addr = None

        return ip_addr

    def validate_ip(self, network_addr, increment=0):
        try:
            ip_addr = str(ipaddress.ip_address(network_addr) + increment)
        except ValueError:
            return None
        return ip_addr


class Subnet:
    def __init__(self, network_addr, cidr):
        self.network = 'SSHA'
        self.sub_start = network_addr
        self.sub_end = ''
        if str(cidr) == '31':
            self.cidr = ''
            try:
                self.sub_end = str(ipaddress.ip_address(network_addr) + 1)
            except ValueError:
                pass
        else:
            self.cidr = cidr
        self.status = 'ACTIVE'
        self.notes = 'SSHA'


class Channel:
    pass


def create_r3_data_file(form):
    # Load a blank template and overwrite old region 3 file
    wb = load_workbook(TEMPLATE_FILE)
    wb.save(R3_FILE)

    rows = {
        'shelf': [],
        'port': [],
        'ip': [],
        'subnet': [],
        'channel': []
    }
    
    d = strip_whitespace(form)
    pe_query = session.query(PeSites).filter(PeSites.shelf_clei == d['peRouter1']).first()
    vpn_concentrators = [pe_query.vpn_primary, pe_query.vpn_backup]
    interfaces = ['physical', 'logical']

    # Shelf rows
    rows['shelf'].append(Shelf(d['cpeRouter'], d['siteCLLI'], d['cpeTemplate'], d['cpeLoopback']))
    # rows['shelf'].update({'ts': Shelf('MINDAO04W', 'MINDAOD00', 'SECURELINX SCL16')})

    # Port rows

    # NMS ports
    rows['port'].extend([
        Port(
            shelf_id=d['peRouter1'],
            device_type='pe',
            path_type='nms',
            interface='logical',
            vlan=d['nmsVlan'],
            network_addr=d['nmsIP']),
        Port(
            shelf_id=d['cpeRouter'],
            device_type='cpe',
            path_type='facility',
            vlan=d['nmsVlan'],
            port=d['rethInterface'],
            bandwidth=1,
            site_id=d['siteCLLI'],
            network_addr=d['nmsIP'])
    ])

    # HOT Facility CPE Port
    rows['port'].append(Port(
        shelf_id=d['cpeRouter'],
        device_type='cpe',
        path_type='facility',
        vlan=d['facilityVlan'],
        port=d['rethInterface'],
        bandwidth=int(d['siteBandwidth']) - 1,
        site_id=d['siteCLLI'],
        network_addr=d['facilityIP']
    ))
    # HOT Facility VPN ports
    for vpn in vpn_concentrators:
        for interface in interfaces:
            rows['port'].append(Port(
                shelf_id=vpn,
                device_type='vpn',
                path_type='facility',
                interface=interface,
                vlan=d['facilityVlan'],
                network_addr=d['facilityIP'],
                bandwidth=int(d['siteBandwidth']) - 1
            ))

    # Service PE ports
    suffix = 1  # Suffix of HTML element names
    for i in range(int(d['numServices'])):
        vlan = '' if suffix == 1 else suffix
        rows['port'].append(Port(
            shelf_id=d['peRouter1'],
            device_type='pe',
            path_type='service',
            interface='logical',
            vlan=d['serviceVlan{}'.format(vlan)],
            network_addr=d['servicePeIP{}'.format(vlan)]
        ))

        for vpn in vpn_concentrators:
            for interface in interfaces:
                rows['port'].append(Port(
                    shelf_id=vpn,
                    device_type='vpn',
                    path_type='service',
                    interface=interface,
                    vlan=d['serviceVlan{}'.format(vlan)],
                    network_addr=d['servicePeIP{}'.format(vlan)]
                ))

        rows['subnet'].extend([
            Subnet(d['serviceCpeIP{}'.format(vlan)], d['serviceCpeCIDR{}'.format(vlan)]),
            Subnet(d['servicePeIP{}'.format(vlan)], d['servicePeCIDR{}'.format(vlan)])
        ])
        suffix += 1

    # Management subnets
    rows['subnet'].extend([
        Subnet(d['nmsIP'], '31'),
        Subnet(d['facilityIP'], '31'),
    ])

    # Append any port rows which have an IP address to IP rows
    for row in rows['port']:
        if row.ip_addr:
            rows['ip'].append(row)

    for sheet, row in rows.items():
        print_sheet(wb, R3_FILE, sheet, row)


def print_sheet(wb, file, sheet, rows):

    sheets = {
        'shelf': wb['SHELF'],
        'port': wb['PORT'],
        'subnet': wb['SUBNET'],
        'ip': wb['IP'],
        'channel': wb['CHANNEL'],
    }

    # Attributes ordered by columns according to Excel template
    sheet_attributes = {
        'shelf': ['shelf_id', 'site_id', 'shelf_template'],
        'port': [
            'shelf_id', 'site_id', 'slot_1',
            'slot_2', 'slot_3', 'port_id',
            'bandwidth', 'connector'],
        'subnet': [
            'network', 'sub_start', 'cidr',
            'sub_end', 'status', 'notes'],
        'ip': [
            'shelf_id', 'site_id', 'slot_1',
            'slot_2', 'slot_3', 'port_id',
            'ip_addr', 'network', 'customer',
            'hostname'],
        'channel': [
            'path', 'channel', 'bandwidth',
            'status']
    }

    active_sheet = sheets[sheet]
    attributes = sheet_attributes[sheet]

    row = 2
    for r in rows:
        col = 1
        for attr in attributes:
            active_sheet['{}{}'.format(get_column_letter(col), row)] = str(getattr(r, attr))
            col += 1
        row += 1

    wb.save(file)


if __name__ == '__main__':
    form = {
        'siteType': 'single',
        'modelName': 'SRX345',
        'numServices': '3',
        'siteCLLI': 'clei',
        'peRouter1': 'TOROONXNPED10',
        'siteBandwidth': '100',
        'siteBandwidthBackup': '',
        'cpeTemplate': 'JUNIPER SRX345 SINGLE CPE',
        'cpeRouter': 'cpe clei',
        'cpeLoopback': '127.0.0.1',
        'rethInterface': 'RETH0',
        'terServer': '',
        'terServerIP1': '',
        'terServerCIDR1': '30',
        'terServerIP2': '',
        'terServerCIDR2': '30',
        'nmsVlan': '600',
        'nmsIP': '0.0.0.0',
        'nmsCIDR': '31',
        'facilityVlan': '601',
        'facilityIP': '20.0.0.0',
        'facilityCIDR': '31',
        'vendorVlan': '',
        'vendorChannel': '',
        'vendorPort': '',
        'nmsVlanBackup': '',
        'nmsIPBackup': '',
        'nmsCIDRBackup': '31',
        'facilityVlanBackup': '',
        'facilityIPBackup': '',
        'facilityCIDRBackup': '31',
        'vendorVlanBackup': '',
        'vendorChannelBackup': '',
        'vendorPortBackup': '',
        'serviceVlan': '100',
        'servicePort': 'GE-0/0/8',
        'serviceCpeIP': '0.0.0.0',
        'serviceCpeCIDR': '30',
        'servicePeIP': '7.0.0.0',
        'servicePeCIDR': '30',
        'serviceVlan2': '101',
        'servicePort2': 'GE-0/0/9',
        'serviceCpeIP2': '1.0.0.0',
        'serviceCpeCIDR2': '30',
        'servicePeIP2': '8.0.0.0',
        'servicePeCIDR2': '30',
        'serviceVlan3': '102',
        'servicePort3': 'GE-0/0/10',
        'serviceCpeIP3': '2.0.0.0',
        'serviceCpeCIDR3': '30',
        'servicePeIP3': '9.0.0.0',
        'servicePeCIDR3': '30',
        'serviceVlanBackup': '',
        'servicePeIPBackup': '',
        'servicePeCIDRBackup': '30'}

    
    create_r3_data_file(form)

    
    pass